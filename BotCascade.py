import sys
import time
import math
import json
import psutil
import requests
import telegram
import pandas as pd
import xlwings as xw
from os import close
import datetime as dt
from datetime import date
from binance.enums import *
from datetime import datetime
from binance.client import Client
from importlib import find_loader
from openpyxl import load_workbook
from win32com.client import DispatchEx
from requests.models import ChunkedEncodingError
from binance.exceptions import BinanceAPIException, BinanceOrderException

path = ""


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Gets Client for Binance API
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def GetClient():
    wb2 = load_workbook(GetPath())
    ws2 = wb2.worksheets[1]

    api_cell = ws2.cell(row = 2, column = 1)
    secret_cell = ws2.cell(row = 2, column = 2)

    api_key =  api_cell.value
    api_secret = secret_cell.value
    
    wb2.close()

    client = Client(api_key, api_secret)
    return client


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Get Balance Futures For Trading
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def CheckBalanceFutures(asset):
    client = GetClient()
    balanceFirst = client.futures_account_balance()
    balanceFirst = next(item for item in balanceFirst if item["asset"] == asset)['balance']
    return balanceFirst


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Check Precision for symbol
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def get_quantity_precision(client,currency_symbol):    
    info = client.futures_exchange_info() 
    info = info['symbols']
    for x in range(len(info)):
        if info[x]['symbol'] == currency_symbol:
            return info[x]['quantityPrecision']
    return None


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Check Balance for specified coin on the Client
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def CheckBalance():
    wb2 = load_workbook(GetPath())
    ws2 = wb2.worksheets[1]

    client = GetClient()

    asset_cell1 = ws2.cell(row = 2, column = 6)
    assetExcel1 = asset_cell1.value

    asset_cell2 = ws2.cell(row = 2, column = 7)
    assetExcel2 = asset_cell2.value

    botType = ws2.cell(row = 2, column = 11).value
    
    if(botType == "Future"):
        balance = client.futures_account_balance()

        balanceFirst = next(item for item in balance if item["asset"] == assetExcel1)['balance']
        balanceSecond = next(item for item in balance if item["asset"] == assetExcel2)['balance']
        
        print ("\n")
        print(assetExcel1,": ", balanceFirst)
        print(assetExcel2,": ", balanceSecond)

        wb2.close()

    elif(botType == "Spot"):
        balanceFirst = client.get_asset_balance(asset=assetExcel1)
    
        print ("\n")
        print(assetExcel1,": ", balanceFirst["free"])

        balanceSecond = client.get_asset_balance(asset=assetExcel2)
        print(assetExcel2,": ", balanceSecond["free"])

        wb2.close()
   


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Hunging orders check for excel refresh
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def AutoCleanCheckCascade(cascadeBuy,cascadeSold,path,cascadeLast,cascadeQuant):
    wbt = xw.Book(path)
    sheet = wbt.sheets[0]

    cascadeBuyCell = sheet.range(str(cascadeBuy)+"3").value
    cascadeSoldCell = sheet.range(str(cascadeSold)+"3").value

    if (cascadeBuyCell == cascadeSoldCell):
        print ("Все ордера в каскаде были завершенны")
        
    if (cascadeBuyCell != cascadeSoldCell):
        buyCellFlag = 2000
        
        while (buyCellFlag >= 6):
            cascadeBuySignal = sheet.range(str(cascadeBuy)+str(buyCellFlag)).value
            
            if (cascadeBuySignal == 1.0):
                wbtNew = xw.Book(GetPath())
                sheetNew = wbtNew.sheets[0]

                for x in range(7,12):
                    excelCopyCheck = sheetNew.range("A"+str(x)).value

                    if (excelCopyCheck == None):
                        copyList = ["A","B","C","D","E","F","G"]

                        for i in copyList:
                            sheetNew.range(i+str(x)).value = sheet.range(i+str(buyCellFlag)).value

                        sheetNew.range(str(cascadeBuy)+str(x)).value = "1"

                        sheetNew.range(str(cascadeQuant)+str(x)).value = sheet.range
                        (str(cascadeQuant)+str(buyCellFlag)).value

                        sheetNew.range(str(cascadeLast)+"1").value = sheet.range
                        (str(cascadeLast)+"1").value

                        wbtNew.save()
                        wbtNew.close()

                        print ("Все ордера в каскаде были скопированны")
                        break

                break
            
            buyCellFlag = buyCellFlag - 1
            
    wbt.close()


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Futures orders check for excel refresh
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def AutoCleanCheckFutures(path,botLinesTransfer,botLinesLimit):
    wbt = xw.Book(path)
    sheet = wbt.sheets[0]

    buyCellFlag = botLinesLimit   
        
    wbtNew = xw.Book(GetPath())
    sheetNew = wbtNew.sheets[0]

    for x in range(7,8+int(botLinesTransfer)):
        buyCellFlag -= 1
        excelCopyCheck = sheetNew.range("A"+str(x)).value

        if (excelCopyCheck == None):
            copyList = ["A","B","C","D","E","F","G","H",
                        "J","T","U","V","W","X","Y","Z",
                        "AA","AB","AC","AE","AF","AG","AH",
                        "AI","AJ","AL","AM","AN","AO","AP",
                        "AQ","AR","AS","AT"]

            for i in copyList:
                sheetNew.range(i+str(x)).value = sheet.range(i+str(buyCellFlag)).value
            
    wbtNew.save()
    wbtNew.close()
    wbt.close()


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Obtain path of the excel file or save it
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def GetPath():
    f = open("excelName.txt","a+")
    f.close()   
    f = open ("excelName.txt","r+")
    contents = f.read()
    
    if (contents == ""):
        print ("Введите имя эксель файла (обязательно добавьте формат файла к концу имени например: .xlsx):\n")
        path = input()
        f.write(path)
        
    else:
        path = contents
        
    f.close()
    return path


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Get path of the excel file template
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def GetPathTemplate():
    f = open("excelNameTemplate.txt","a+")
    f.close()   
    f = open ("excelNameTemplate.txt","r+")
    contents = f.read()
    
    if (contents == ""):
        print ("Введите имя шаблон файла (обязательно добавьте формат файла к концу имени например: .xlsx):\n")
        path = input()
        f.write(path)
        
    else:
        path = contents
        
    f.close()
    return path


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Fire order buy FUTURE
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def BuyOrderFuture(allocFunds,btc_price,client,coin,
             ws2,ws1,flag,assetExcel1,assetExcel2,
             wb2,buyFlag,quantFlag,path,quantBuyFlag):
    #TODO add futures option
    while (buyFlag > 0):
        try:
            quant = float(allocFunds)/float(btc_price["price"])
            quant = float(round(quant,get_quantity_precision(client,coin)))

            orderType = ws1.cell(row = 2, column = 10)
            
            # if (orderType.value == "LIMIT"):
            #     buy_order_limit = client.futures_create_order(
            #         symbol=coin,
            #         side='BUY',
            #         type='LIMIT',
            #         timeInForce='GTC',
            #         quantity=final,
            #         price = int(c6))

            if (orderType.value == "MARKET"):
                buy_order_limit = client.futures_create_order(
                    symbol=coin,
                    side='BUY',
                    positionSide = botHedgeType,
                    type='MARKET',
                    quantity=quant)

            print("\n\n-------------------------------------------------")
            print (datetime.now(),"Ордер на покупку был выставлен:")
            print("\n",buy_order_limit)
            print("-------------------------------------------------")

            time.sleep(5)
            order_confirm = client.futures_get_order(
                        symbol = coin,
                        orderId = buy_order_limit["orderId"]
                    )
            
            if (order_confirm["status"]== "FILLED"): 
                cBoughtQuant = ws2.cell(row = flag, column = quantBuyFlag)
                cBoughtQuant.value = float(btc_price["price"]) * float(order_confirm["executedQty"])

                # cBoughtPrice = ws2.cell(row = flag, column = 10)
                # cBoughtPrice.value = float(btc_price["price"])

                сFirstAssetBalance = ws1.cell(row = 2, column = 8)
                сFirstAssetBalance.value = CheckBalanceFutures(assetExcel1)

                cSecondAssetBalance = ws1.cell(row = 2, column = 9)
                cSecondAssetBalance.value = CheckBalanceFutures(assetExcel2)

                # cBoughtLastQuant = ws2.cell(row = 1, column = quantFlag)
                # cBoughtLastQuant.value = buy_order_limit["executedQty"]

                wb2.save(path)
                
                print("\n\n-------------------------------------------------")
                print(datetime.now(),"покупка прошла успешно, данные о покупке:")
                print(order_confirm)

                buyFlag = 10
                time.sleep(5) 
                    
                TelegramBotOrder("Произошла покупка фьючерса, Кол-во: "
                                    +str(cBoughtQuant.value)+"; Цена: "
                                    +str(btc_price["price"])+"; Баланс "
                                    +str(assetExcel1)+": "
                                    +str(сFirstAssetBalance.value)+"; Баланс "
                                    +str(assetExcel2)+": "
                                    +str(cSecondAssetBalance.value),path)
                
            else:
                client.futures_cancel_order(symbol=coin, orderId=buy_order_limit["orderId"])

                print("Ордер не был заполнен, ордер был отменен, данные были удаленны")

                c1 = ws2.cell(row = flag, column = 28)
                c1.value = None

                wb2.save(path)
                
            break
            
        except BinanceAPIException as e:      
            print(e)
            print(datetime.now(),
                  "Произошла ошибка во время подключения, количество оставшихся попыток: "
                  + str(buyFlag))

            TelegramBot("Произошла ошибка во время подключения с бинансом количество оставшихся попыток:"
                        + str(buyFlag)+ str(e),path)

            if (buyFlag > 0):
                print ("Операция повторится через 5 секунд")
                time.sleep(5) 

                if (buyFlag == 1):
                    print ("Бот не смог подключиться данные о покупке были удаленны")

                    c1 = ws2.cell(row = flag, column = 28)
                    c1.value = None

                    wb2.save(path)

                    StartBot(path)

                buyFlag = buyFlag - 1

            else:
                StartBot(path)
                
        except BinanceOrderException as e:
            # error handling goes here
            print(datetime.now(),
                  "Произошла ошибка во время проведения операции с бинансом сохраните этот код ошибки:")

            TelegramBot("Произошла ошибка во время проведения операции с бинансом бот был остановлен. Код ошибки:"
                        + str(e),path)

            print(e)
            Menu()
            
        except Exception as e:
            print(datetime.now(),"Нету связи с интернетом бот попробует записать данные еще раз через 1 минуту "
                  + str(e))
            time.sleep(60) 


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Fire sell order FUTURE
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def SellOrderFuture(assetQuant,sellFlag,client,coin,ws1,ws2,
              flag,assetExcel1,assetExcel2,wb2,path,
              btc_price,quantSellFlag):
    #TODO Add futures option
    while (sellFlag > 0):                       
        try:
            quant = float(assetQuant)/float(btc_price["price"])
            quant = float(round(quant,get_quantity_precision(client,coin)))

            orderType = ws1.cell(row = 2, column = 10)

            # if (orderType.value == "LIMIT"):
            #     sell_order_limit = client.create_order(
            #         symbol=coin,
            #         side='SELL',
            #         type='LIMIT',
            #         timeInForce='GTC',
            #         quantity=final,
            #         price = int(c7))
            
            if (orderType.value == "MARKET"):
                sell_order_limit = client.futures_create_order(
                    symbol=coin,
                    side='SELL',
                    positionSide = botHedgeType,
                    type='MARKET',
                    quantity=quant)

            print("\n\n-------------------------------------------------")
            print (datetime.now(),"Ордер на продажу был выставлен:")
            print("\n",sell_order_limit)
            print("-------------------------------------------------")

            time.sleep(5)
            
            order_confirm = client.futures_get_order(
                        symbol = coin,
                        orderId = sell_order_limit["orderId"]
                    )
            
            if (order_confirm["status"]== "FILLED"): 

                cBoughtQuant = ws2.cell(row = flag, column = quantSellFlag)
                cBoughtQuant.value = float(btc_price["price"]) * float(order_confirm["executedQty"])

                # cBoughtPrice = ws2.cell(row = flag, column = 10)
                # cBoughtPrice.value = float(btc_price["price"])

                сFirstAssetBalance = ws1.cell(row = 2, column = 8)
                сFirstAssetBalance.value = CheckBalanceFutures(assetExcel1)

                cSecondAssetBalance = ws1.cell(row = 2, column = 9)
                cSecondAssetBalance.value = CheckBalanceFutures(assetExcel2)

                wb2.save(path)

                print("\n\n-------------------------------------------------")
                print(datetime.now(),"продажа прошла успешно, данные о продаже:")
                print(order_confirm)
                
                sellFlag = 10
                time.sleep(5)
                
                TelegramBotOrder("Произошла продажа, Кол-во: "
                                    +str(cBoughtQuant.value)+"; Цена: "
                                    +str(btc_price["price"])+"; Баланс "
                                    +str(assetExcel1)+": "
                                    +str(сFirstAssetBalance.value)+"; Баланс "
                                    +str(assetExcel2)+": "
                                    +str(cSecondAssetBalance.value),path)
                
                 
                
            else:
                client.futures_cancel_order(symbol=coin, orderId=sell_order_limit["orderId"])

                print("Ордер не был заполнен, ордер был отменен, данные были удаленны")

                c1 = ws2.cell(row = flag, column = 25)
                c1.value = None
                
                wb2.save(path)

            break
        
        except BinanceAPIException as e:
            # error handling goes here
            print(datetime.now(),"Произошла ошибка во время подключения, количество оставшихся попыток: "
                + str(sellFlag))

            TelegramBot("Произошла ошибка во время подключения с бинансом количество оставшихся попыток:" 
                + str(sellFlag)+ str(e),path)

            print(e)
            
            if (sellFlag > 0):
                print ("Операция повторится через 5 секунд")
                time.sleep(5) 
                
                if (sellFlag == 1):
                    print ("Бот не смог подключиться данные о продаже были удаленны")

                    c1 = ws2.cell(row = flag, column = 25)
                    c1.value = None

                    wb2.save(path)
                    StartBot(path)

                sellFlag = sellFlag - 1

            else:
                StartBot(path)
                
        except BinanceOrderException as e:
            # error handling goes here
            print(datetime.now(),
                "Произошла ошибка во время проведения операции с бинансом сохраните этот код ошибки:")

            TelegramBot("Произошла ошибка во время проведения операции с бинансом бот был остановлен. Код ошибки:"
                + str(e),path)

            print(e)
            Menu()
            
        except Exception as e:
            print(datetime.now(),"Нету связи с интернетом бот попробует записать данные еще раз через 1 минуту "
                + str(e))
            time.sleep(60) 


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Fire order buy
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def BuyOrder(allocFunds,btc_price,client,coin,c6,
             ws2,ws1,flag,assetExcel1,assetExcel2,
             wb2,buyFlag,quantFlag,path,buyOrderList,
             statusFlag,quantBuyFlag):
    #TODO add futures option
    while (buyFlag > 0):
        try:
            if (statusFlag == 0):
                quant = float(allocFunds)/float(btc_price["price"])

                info = client.get_symbol_info(coin)
                step_size = [float(_['stepSize']) for _ in info['filters'] if _['filterType'] == 'LOT_SIZE'][0]
                step_size = '%.8f' % step_size
                step_size = step_size.rstrip('0')
                decimals = len(step_size.split('.')[1])
                final = math.floor(quant * 10.0001 ** decimals) / 10 ** decimals
                
                orderType = ws1.cell(row = 2, column = 10)
                
                if (orderType.value == "LIMIT"):
                    buy_order_limit = client.create_order(
                        symbol=coin,
                        side='BUY',
                        type='LIMIT',
                        timeInForce='GTC',
                        quantity=final,
                        price = int(c6))

                elif (orderType.value == "MARKET"):
                    buy_order_limit = client.create_order(
                        symbol=coin,
                        side='BUY',
                        type='MARKET',
                        quantity=final)

                print("\n\n-------------------------------------------------")
                print (datetime.now(),"Ордер на покупку был выставлен:")
                print("\n",buy_order_limit)
                print("-------------------------------------------------")

                buyOrderList.append(buy_order_limit["orderId"])

                buyFlag = 10
                break

            if (statusFlag == 1):
                for i in buyOrderList:
                    order_confirm = client.get_order(
                        symbol = coin,
                        orderId = i
                    )
                    
                    if (order_confirm["status"]== "FILLED"):
                        cBoughtQuant = ws2.cell(row = flag, column = quantBuyFlag)
                        cBoughtQuant.value = float(btc_price["price"]) * float(order_confirm["executedQty"])

                        cBoughtPrice = ws2.cell(row = flag, column = 10)
                        cBoughtPrice.value = float(btc_price["price"])

                        balanceFirst = client.get_asset_balance(asset=assetExcel1)
                        сFirstAssetBalance = ws1.cell(row = 2, column = 8)
                        сFirstAssetBalance.value = balanceFirst["free"]

                        balanceSecond = client.get_asset_balance(asset=assetExcel2)
                        cSecondAssetBalance = ws1.cell(row = 2, column = 9)
                        cSecondAssetBalance.value = balanceSecond["free"]

                        cBoughtLastQuant = ws2.cell(row = 1, column = quantFlag)
                        cBoughtLastQuant.value = order_confirm["executedQty"]

                        wb2.save(path)
                        
                        print("\n\n-------------------------------------------------")
                        print(datetime.now(),"покупка прошла успешно, данные о покупке:")
                        print(order_confirm)

                        buyFlag = 10
                        time.sleep(15) 
                        
                        cascade = 0

                        if (quantBuyFlag== 17):
                            cascade = 1
                        elif(quantBuyFlag==34): 
                            cascade = 2
                        elif(quantBuyFlag==51): 
                            cascade = 3  
                        elif(quantBuyFlag==68): 
                            cascade = 4
                        elif(quantBuyFlag==85): 
                            cascade = 5
                            
                        TelegramBotOrder("Произошла покупка, Каскад: "
                                         +str(cascade)+" Кол-во: "
                                         +str(cBoughtQuant.value)+"; Цена: "
                                         +str(cBoughtPrice.value)+"; Баланс "
                                         +str(assetExcel1)+": "
                                         +str(сFirstAssetBalance.value)+"; Баланс "
                                         +str(assetExcel2)+": "
                                         +str(cSecondAssetBalance.value),path)
                        
                    else:
                        client.cancel_order(symbol=coin, orderId=i)

                        print("Ордер не был заполнен, ордер был отменен, данные были удаленны")

                        c1 = ws2.cell(row = flag, column = 9)
                        c1.value = None

                        c2 = ws2.cell(row = flag, column = 10)
                        c2.value = None

                        c3 = ws2.cell(row = flag, column = 18)
                        c3.value = None

                        c4 = ws2.cell(row = flag, column = 35)
                        c4.value = None

                        c5 = ws2.cell(row = flag, column = 52)
                        c5.value = None
                        
                        c6 = ws2.cell(row = flag, column = 69)
                        c6.value = None

                        c7 = ws2.cell(row = flag, column = 86)
                        c7.value = None

                        wb2.save(path)
                        
                break
            
        except BinanceAPIException as e:      
            print(e)
            print(datetime.now(),
                  "Произошла ошибка во время подключения, количество оставшихся попыток: "
                  + str(buyFlag))

            TelegramBot("Произошла ошибка во время подключения с бинансом количество оставшихся попыток:"
                        + str(buyFlag)+ str(e),path)

            if (buyFlag > 0):
                print ("Операция повторится через 5 секунд")
                time.sleep(5) 

                if (buyFlag == 1):
                    print ("Бот не смог подключиться данные о покупке были удаленны")

                    c1 = ws2.cell(row = flag, column = 9)
                    c1.value = None

                    c2 = ws2.cell(row = flag, column = 10)
                    c2.value = None

                    c3 = ws2.cell(row = flag, column = 18)
                    c3.value = None

                    c4 = ws2.cell(row = flag, column = 35)
                    c4.value = None

                    c5 = ws2.cell(row = flag, column = 52)
                    c5.value = None
                    
                    c6 = ws2.cell(row = flag, column = 69)
                    c6.value = None

                    c7 = ws2.cell(row = flag, column = 86)
                    c7.value = None

                    wb2.save(path)

                    StartBot(path)

                buyFlag = buyFlag - 1

            else:
                StartBot(path)
                
        except BinanceOrderException as e:
            # error handling goes here
            print(datetime.now(),
                  "Произошла ошибка во время проведения операции с бинансом сохраните этот код ошибки:")

            TelegramBot("Произошла ошибка во время проведения операции с бинансом бот был остановлен. Код ошибки:"
                        + str(e),path)

            print(e)
            Menu()
            
        except Exception as e:
            print(datetime.now(),"Нету связи с интернетом бот попробует записать данные еще раз через 1 минуту "
                  + str(e))
            time.sleep(60) 


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Fire sell order
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def SellOrder(sellFlag,assetQuant,client,coin,c7,ws1,ws2,
              flag,assetExcel1,assetExcel2,wb2,path,
              sellOrderList,statusFlag,sellOrderDict,
              btc_price,signalSellFlag):
    #TODO Add futures option
    while (sellFlag > 0):            
                            
        try:
            if (statusFlag == 0):
                info = client.get_symbol_info(coin)
                step_size = [float(_['stepSize']) 
                             for _ in info['filters'] 
                             if _['filterType'] == 'LOT_SIZE'][0]
                step_size = '%.8f' % step_size
                step_size = step_size.rstrip('0')
                decimals = len(step_size.split('.')[1])
                final = math.floor(assetQuant * 10.0001 ** decimals) / 10 ** decimals

                orderType = ws1.cell(row = 2, column = 10)

                if (orderType.value == "LIMIT"):
                    sell_order_limit = client.create_order(
                        symbol=coin,
                        side='SELL',
                        type='LIMIT',
                        timeInForce='GTC',
                        quantity=final,
                        price = int(c7))
                
                elif (orderType.value == "MARKET"):
                    sell_order_limit = client.create_order(
                        symbol=coin,
                        side='SELL',
                        type='MARKET',
                        quantity=final)

                print("\n\n-------------------------------------------------")
                print (datetime.now(),"Ордер на продажу был выставлен:")
                print("\n",sell_order_limit)
                print("-------------------------------------------------")

                sellOrderList.append(sell_order_limit["orderId"])
                sellFlag = 10

                break

            elif (statusFlag == 1):
                dict_pairs = sellOrderDict.items()
                pairs_iterator = iter(dict_pairs)
                
                for i in sellOrderList:
                    order_confirm = client.get_order(
                        symbol = coin,
                        orderId = i
                    )

                    if (order_confirm["status"]== "FILLED"):
                        SellPair = next(pairs_iterator)

                        cBoughtQuant = ws2.cell(row = flag, column = SellPair[0])
                        cBoughtQuant.value = float(btc_price["price"]) * float(order_confirm["executedQty"])

                        cBoughtPrice = ws2.cell(row = flag, column = 10)
                        cBoughtPrice.value = float(btc_price["price"])

                        balanceFirst = client.get_asset_balance(asset=assetExcel1)
                        сFirstAssetBalance = ws1.cell(row = 2, column = 8)
                        сFirstAssetBalance.value = balanceFirst["free"]

                        balanceSecond = client.get_asset_balance(asset=assetExcel2)
                        cSecondAssetBalance = ws1.cell(row = 2, column = 9)
                        cSecondAssetBalance.value = balanceSecond["free"]

                        wb2.save(path)

                        print("\n\n-------------------------------------------------")
                        print(datetime.now(),"продажа прошла успешно, данные о продаже:")
                        print(order_confirm)
                        
                        cascade = 0

                        if (SellPair[0]== 20):
                            cascade = 1
                        elif(SellPair[0]==37): 
                            cascade = 2
                        elif(SellPair[0]==54): 
                            cascade = 3  
                        elif(SellPair[0]==71): 
                            cascade = 4
                        elif(SellPair[0]==88): 
                            cascade = 5

                        TelegramBotOrder("Произошла продажа, Каскад: "
                                         +str(cascade)+" Кол-во: "
                                         +str(cBoughtQuant.value)+"; Цена: "
                                         +str(cBoughtPrice.value)+"; Баланс "
                                         +str(assetExcel1)+": "
                                         +str(сFirstAssetBalance.value)+"; Баланс "
                                         +str(assetExcel2)+": "
                                         +str(cSecondAssetBalance.value),path)
                        
                        sellFlag = 10
                        time.sleep(15) 
                        
                    else:
                        SellPair = next(pairs_iterator)

                        client.cancel_order(symbol=coin, orderId=i)

                        print("Ордер не был заполнен, ордер был отменен, данные были удаленны")

                        c1 = ws2.cell(row = flag, column = 9)
                        c1.value = None

                        c2 = ws2.cell(row = flag, column = 10)
                        c2.value = None

                        c3 = ws2.cell(row = flag, column = SellPair[1])
                        c3.value = None
                    
                        wb2.save(path)

                break
        
        except BinanceAPIException as e:
            # error handling goes here
            print(datetime.now(),"Произошла ошибка во время подключения, количество оставшихся попыток: "
                + str(sellFlag))

            TelegramBot("Произошла ошибка во время подключения с бинансом количество оставшихся попыток:" 
                + str(sellFlag)+ str(e),path)

            print(e)
            
            if (sellFlag > 0):
                print ("Операция повторится через 5 секунд")
                time.sleep(5) 
                
                if (sellFlag == 1):
                    print ("Бот не смог подключиться данные о продаже были удаленны")

                    c1 = ws2.cell(row = flag, column = signalSellFlag)
                    c1.value = None

                    wb2.save(path)
                    StartBot(path)

                sellFlag = sellFlag - 1

            else:
                StartBot(path)
                
        except BinanceOrderException as e:
            # error handling goes here
            print(datetime.now(),
                "Произошла ошибка во время проведения операции с бинансом сохраните этот код ошибки:")

            TelegramBot("Произошла ошибка во время проведения операции с бинансом бот был остановлен. Код ошибки:"
                + str(e),path)

            print(e)
            Menu()
            
        except Exception as e:
            print(datetime.now(),"Нету связи с интернетом бот попробует записать данные еще раз через 1 минуту "
                + str(e))
            time.sleep(60) 


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Sell specified amount of specified coin on Client
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def SellFunc(path):
    try:
        wb2 = load_workbook(path)
        ws2 = wb2.worksheets[1]
        ws1 = wb2.worksheets[0]
        wb2.save(path)
        
    except Exception as e:
        print("\n-------------------------------------------------")
        print("Введенное вами имя эксель файла не найденно попробуйте еще раз")
        print("-------------------------------------------------")
        print(e)
        Menu()

    client = GetClient()

    asset_cell1 = ws2.cell(row = 2, column = 6)
    assetExcel1 = asset_cell1.value

    asset_cell2 = ws2.cell(row = 2, column = 7)
    assetExcel2 = asset_cell2.value

    coin_cell = ws2.cell(row = 2, column = 5)
    coin = coin_cell.value

    coin_test = client.get_symbol_ticker(symbol=coin)

    asset_test1 = client.get_asset_balance(asset=assetExcel1)

    asset_test2 = client.get_asset_balance(asset=assetExcel2)

    balanceFirst = client.get_asset_balance(asset=assetExcel1)

    print ("\nБаланс указанной вами валюты (",str(assetExcel1),"): ",balanceFirst["free"])                           
    print("Введите количество валюты в процентах которое вы желаете продать(от 10 до 100 знак '%' не нужен): \n")

    totalToSell = input()

    try:
        totalToSell = int(totalToSell)

    except Exception as e:
        print("\n-------------------------------------------------")
        print("Введенный вами процент не соблюдает формат попробуйте еще раз")
        print("-------------------------------------------------")
        print(e)
        SellFunc(path) 

    if (int(totalToSell) > 9 and int(totalToSell) < 101):
        sellFlag = 10

        calcQuant = (int(totalToSell) / 100) * float(balanceFirst["free"])

        while (sellFlag > 0):                                
            try:
                info = client.get_symbol_info(coin)
                step_size = [float(_['stepSize']) for _ in info['filters'] if _['filterType'] == 'LOT_SIZE'][0]
                step_size = '%.8f' % step_size
                step_size = step_size.rstrip('0')
                decimals = len(step_size.split('.')[1])
                final = math.floor(calcQuant * 10.0001 ** decimals) / 10 ** decimals

                sell_order_limit = client.create_order(
                    symbol=coin,
                    side='SELL',
                    type='MARKET',
                    quantity=final)
                
                print("\n\n-------------------------------------------------")
                print(datetime.now(),"Произошла продажа, данные о продаже:")
                print(sell_order_limit)

                balanceFirst = client.get_asset_balance(asset=assetExcel1)
                сFirstAssetBalance = ws2.cell(row = 2, column = 8)
                сFirstAssetBalance.value = balanceFirst["free"]

                balanceSecond = client.get_asset_balance(asset=assetExcel2)
                cSecondAssetBalance = ws2.cell(row = 2, column = 9)
                cSecondAssetBalance.value = balanceSecond["free"]

                wb2.save(path)

                sellFlag = 10

                Menu()
                break
            
            except BinanceAPIException as e:
                # error handling goes here
                print(datetime.now(),"Произошла ошибка во время подключения, количество оставшихся попыток: "
                    + str(sellFlag))
                print(e)
                
                if (sellFlag > 0):
                    sellFlag = sellFlag - 1

                    print ("Операция повторится через 5 секунд")
                    time.sleep(5) 

                    if (sellFlag == 0):
                            Menu()

                else:
                    Menu()
                    
            except BinanceOrderException as e:
                # error handling goes here
                print(datetime.now(),
                    "Произошла ошибка во время проведения операции с бинансом сохраните этот код ошибки:")
                print(e)
                Menu()


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Refresh excel file if limit of cells is reached
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def ClearExcel():
    newPath = GetPathTemplate()

    try:
        wb2 = load_workbook(newPath)
        newPath = str(datetime.now().strftime("%Y-%m-%d"))+newPath
        wb2.save(newPath)
        wb2.close()
        print ("Файл готов торговля начинается\n")

        f = open("excelName.txt","w+")
        f.write(newPath)
        f.close()   

    except Exception as e:
        print("\n-------------------------------------------------")
        print("Введенное вами имя эксель файла не найденно попробуйте еще раз")
        print("-------------------------------------------------")
        print(e)
        Menu()


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Body of the bot
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#TODO simplify the logic decrease complexity
def MainFunc(path):
    try:
        wbt2 = xw.Book("dontTouch.xlsx")
        wbt2.app.visible = False
        
    except Exception as e:
        if ("No such file" in str(e)):
            wbt2 = xw.Book()
            wbt2.save("dontTouch.xlsx")
            wbt2.close()
            MainFunc(path)
        print ("Возникла ошибка: ",e)

    try:
        wb2 = load_workbook(path)
        ws2 = wb2.worksheets[1]
        ws1 = wb2.worksheets[0]
        wb2.save(path)
        
    except Exception as e:
        print("\n-------------------------------------------------")
        print("Введенное вами имя эксель файла не найденно попробуйте еще раз")
        print("-------------------------------------------------")
        print(e)
        Menu()

    client = GetClient()

    time_cell = ws2.cell(row = 2, column = 4)
    orderTime = time_cell.value
    timesList = []

    funds_cell = ws2.cell(row = 2, column = 3)

    coin_cell = ws2.cell(row = 2, column = 5)
    coin = coin_cell.value

    asset_cell1 = ws2.cell(row = 2, column = 6)
    assetExcel1 = asset_cell1.value

    asset_cell2 = ws2.cell(row = 2, column = 7)
    assetExcel2 = asset_cell2.value
    
    global botType 
    botType = ws2.cell(row = 2, column = 11).value
        
    global botHedgeType 
    botHedgeType = ws2.cell(row = 2, column = 19).value
    
    botLineLimit = ws2.cell(row = 2, column = 17).value
    botLinesTransfer = ws2.cell(row = 2, column = 18).value

    if (botType == "Future"):
        botLeverage = ws2.cell(row = 2, column = 13).value
        client.futures_change_leverage(symbol=coin, leverage=int(botLeverage))
        
        botPositionMode = ws2.cell(row = 2, column = 15).value
        if (botPositionMode == "HEDGE" and client.futures_get_position_mode()["dualSidePosition"] != True):            
            client.futures_change_position_mode(dualSidePosition="true")
        elif (botPositionMode == "ONEWAY" and client.futures_get_position_mode()["dualSidePosition"] != False):
            client.futures_change_position_mode(dualSidePosition="false")
            
        botAssetMode = ws2.cell(row = 2, column = 16).value
        if (botAssetMode == "MULTI" and client.futures_get_multi_assets_mode()["multiAssetsMargin"] != True):
            client.futures_change_multi_assets_mode(multiAssetsMargin="true")
        elif (botAssetMode == "SINGLE" and client.futures_get_multi_assets_mode()["multiAssetsMargin"] != False):
            client.futures_change_multi_assets_mode(multiAssetsMargin="false")

    asset_test1 = client.get_asset_balance(asset=assetExcel1)

    asset_test2 = client.get_asset_balance(asset=assetExcel2)

    coin_test = client.get_symbol_ticker(symbol=coin)

    global symbolPrecision
    symbolPrecision = get_quantity_precision(client,coin)

    try:
        allocFunds = int(funds_cell.value)
        if (allocFunds < 10):
            print("\n-------------------------------------------------")
            print("Минимальная сумма для торговли: (10$) введенная вами сумма меньше, попробуйте еще раз")
            print("-------------------------------------------------")
            Menu()

    except Exception as e:
        print("\n-------------------------------------------------")
        print("Введенное вами сумма для торговли не соблюдает формат попробуйте еще раз")
        print("-------------------------------------------------")
        print(e)
        Menu()

    if (str(orderTime) == "1"):
        timesList = ['00','01','02','03','04','05','06','07',
                     '08','09','10','11','12','13','14','15',
                     '16','17','18','19','20','21','22','23',
                     '24','25','26','27','28','29','30','31',
                     '32','33','34','35','36','37','38','39',
                     '40','41','42','43','44','45','46','47',
                     '48','49','50','51','52','53','54','55',
                     '56','57','58','59']
    
    elif (str(orderTime) == "5"):
        timesList = ["00","05","10","15","20","25","30","35","40","45","50","55"]
   
    elif (str(orderTime) == "15"):
        timesList = ["00","15","30","45"]
    
    elif (str(orderTime) == "30"):
        timesList = ['00','30']

    else:
        print ("\nВремя указанное в экселе не соблюдает формату проверьте свой файл")
        Menu()

    flag = 7
    buyFlag = 10
    sellFlag = 10
    
    if (botType == "Future"):
        сFirstAssetBalance = ws2.cell(row = 2, column = 8)
        сFirstAssetBalance.value = CheckBalanceFutures(assetExcel1)

        cSecondAssetBalance = ws2.cell(row = 2, column = 9)
        cSecondAssetBalance.value = CheckBalanceFutures(assetExcel2)
        
    elif (botType == "Spot"):
        balanceFirst = client.get_asset_balance(asset=assetExcel1)
        сFirstAssetBalance = ws2.cell(row = 2, column = 8)
        сFirstAssetBalance.value = balanceFirst["free"]

        balanceSecond = client.get_asset_balance(asset=assetExcel2)
        cSecondAssetBalance = ws2.cell(row = 2, column = 9)
        cSecondAssetBalance.value = balanceSecond["free"]
    
    wb2.save(path)
    wb2.close()
    while True:
        client = GetClient()
        wb2 = load_workbook(path)
        ws2 = wb2.worksheets[0]
        ws1 = wb2.worksheets[1]
        c1 = ws2.cell(row = 1, column = 1)
        now = datetime.now()
        currentTime = now.strftime("%M")
        
        if (currentTime in timesList):
            while (flag < 1048575):
                c1 = ws2.cell(row = flag, column = 1)
                if (flag >= int(botLineLimit)):

                    oldPath = GetPath()

                    ClearExcel()
                    
                    if(botType == "Future"):
                        AutoCleanCheckFutures(oldPath,botLinesTransfer,botLineLimit)
                    else:
                        AutoCleanCheckCascade("R","U",oldPath,"N","Q")
                        AutoCleanCheckCascade("AI","AL",oldPath,"AE","AH")
                        AutoCleanCheckCascade("AZ","BC",oldPath,"AV","AY")
                        AutoCleanCheckCascade("BQ","BT",oldPath,"BM","BP")
                        AutoCleanCheckCascade("CH","CK",oldPath,"CD","CG")

                    StartBot(GetPath())

                if (c1.value == None):
                    if (botType == "Future"):
                        btc_price = client.futures_symbol_ticker(symbol=coin)
                        hist_klines = client.futures_klines(symbol=coin, interval=str(orderTime)+"m",limit=1)
                        btc_price_high = hist_klines[0][2]
                        btc_price_low = hist_klines[0][3]
                        btc_price_open =hist_klines[0][1]

                    elif (botType == "Spot"):
                        btc_price = client.get_symbol_ticker(symbol=coin)
                        hist_klines = client.get_klines(symbol=coin, interval=str(orderTime)+"m", limit= 1)
                        btc_price_high = hist_klines[0][2]
                        btc_price_low = hist_klines[0][3]
                        btc_price_open =hist_klines[0][1]

                    now = datetime.now()
                    currentTime = now.strftime("%H:%M")
                    
                    c1.value = date.today()
                
                    c2 = ws2.cell(row = flag, column = 2)
                    c2.value = currentTime

                    orderType = ws1.cell(row = 2, column = 10)

                    c3 = ws2.cell(row = flag, column = 6)
                    c3.value = float(btc_price["price"])
                    c3.number_format

                    c4_high = ws2.cell(row = flag, column = 4)
                    c4_high.value = float(btc_price_high)
                    c4_high.number_format

                    c5_low = ws2.cell(row = flag, column = 5)
                    c5_low.value = float(btc_price_low)
                    c5_low.number_format

                    c6_open = ws2.cell(row = flag, column = 3)
                    c6_open.value = float(btc_price_open)
                    c6_open.number_format

                    wb2.save(path)
                    
                    wbt = xw.Book(path)
                    sheet = wbt.sheets[0]
                    flagstr = str(flag)
                    
                    #TODO add futures bot
                    if (botType == "Future"):
                        buySignalFuture = sheet.range('AL'+flagstr).value
                        sellSignalFuture = sheet.range('AF'+flagstr).value
                        
                        if(buySignalFuture == 1 or sellSignalFuture == 1):
                            if (buySignalFuture == 1):
                                wbt = xw.Book(path)
                                sheet = wbt.sheets[0]
                                quantToBuy = sheet.range('AJ'+flagstr).value

                                print(float(btc_price["price"]))
                                wbt.close()

                                quantBuyFlag = 37
                                quantFlag = 37

                                BuyOrderFuture(quantToBuy,btc_price,client,coin,ws2,
                                        ws1,flag,assetExcel1,assetExcel2,wb2,
                                        buyFlag,quantFlag,path,
                                        quantBuyFlag)
                                
                            if (sellSignalFuture == 1):
                                wbt = xw.Book(path)
                                sheet = wbt.sheets[0]
                                quantToSell = sheet.range('AI'+flagstr).value

                                print(float(btc_price["price"]))
                                wbt.close()
                                
                                quantSellFlag = 30

                                SellOrderFuture(quantToSell,sellFlag,client,coin,ws1,ws2,
                                        flag,assetExcel1,assetExcel2,wb2,path
                                        ,btc_price,quantSellFlag)
                            break
                            
                        else:
                            print("\n-------------------------------------------------")
                            print (datetime.now(),"Данные были введенны в таблицу, никакого действия не было обнаруженно, бот продолжит работать")
                            print("-------------------------------------------------")
                            wbt.close()
                            break
                        
                    elif (botType == "Spot"):
                        buySignal1 = sheet.range('R'+flagstr).value
                        buySignal2 = sheet.range('AI'+flagstr).value
                        buySignal3 = sheet.range('AZ'+flagstr).value
                        buySignal4 = sheet.range('BQ'+flagstr).value
                        buySignal5 = sheet.range('CH'+flagstr).value

                        sellSignal1 = sheet.range('U'+flagstr).value
                        sellSignal2 = sheet.range('AL'+flagstr).value
                        sellSignal3 = sheet.range('BC'+flagstr).value
                        sellSignal4 = sheet.range('BT'+flagstr).value
                        sellSignal5 = sheet.range('CK'+flagstr).value

                        loopFlag = 0
                        outerStatusFlag = 0
                        innerStatusFlag = 0
                        
                        buyOrderList = []
                        sellOrderList = []
                        sellOrderDict = {}
                    
                        if (buySignal1 == 1):
                            wbt = xw.Book(path)
                            sheet = wbt.sheets[0]
                            c6 = sheet.range('S'+flagstr).value
                            print(c6)
                            wbt.close()

                            quantBuyFlag = 17
                            quantFlag = 14

                            BuyOrder(allocFunds,btc_price,client,coin,c6,ws2,
                                    ws1,flag,assetExcel1,assetExcel2,wb2,
                                    buyFlag,quantFlag,path,buyOrderList,
                                    outerStatusFlag,quantBuyFlag)

                            loopFlag = 1
                            innerStatusFlag = 1
                        
                        if (buySignal2 == 1):
                            wbt = xw.Book(path)
                            sheet = wbt.sheets[0]
                            c6 = sheet.range('AJ'+flagstr).value
                            print(c6)
                            wbt.close()

                            quantBuyFlag = 34
                            quantFlag = 31

                            BuyOrder(allocFunds,btc_price,client,coin,c6,ws2,
                                    ws1,flag,assetExcel1,assetExcel2,wb2,
                                    buyFlag,quantFlag,path,buyOrderList,
                                    outerStatusFlag,quantBuyFlag)

                            loopFlag = 1
                            innerStatusFlag = 1

                        if (buySignal3 == 1):
                            wbt = xw.Book(path)
                            sheet = wbt.sheets[0]
                            c6 = sheet.range('BA'+flagstr).value
                            print(c6)
                            wbt.close()

                            quantBuyFlag = 51
                            quantFlag = 48

                            BuyOrder(allocFunds,btc_price,client,coin,c6,ws2,
                                    ws1,flag,assetExcel1,assetExcel2,wb2,
                                    buyFlag,quantFlag,path,buyOrderList,
                                    outerStatusFlag,quantBuyFlag)

                            loopFlag = 1
                            innerStatusFlag = 1

                        if (buySignal4 == 1):
                            wbt = xw.Book(path)
                            sheet = wbt.sheets[0]
                            c6 = sheet.range('BR'+flagstr).value
                            print(c6)
                            wbt.close()

                            quantBuyFlag = 68
                            quantFlag = 65

                            BuyOrder(allocFunds,btc_price,client,coin,c6,ws2,
                                    ws1,flag,assetExcel1,assetExcel2,wb2,
                                    buyFlag,quantFlag,path,buyOrderList,
                                    outerStatusFlag,quantBuyFlag)

                            loopFlag = 1
                            innerStatusFlag = 1

                        if (buySignal5 == 1):
                            wbt = xw.Book(path)
                            sheet = wbt.sheets[0]
                            c6 = sheet.range('CI'+flagstr).value
                            print(c6)
                            wbt.close()

                            quantBuyFlag = 85
                            quantFlag = 82

                            BuyOrder(allocFunds,btc_price,client,coin,c6,ws2,
                                    ws1,flag,assetExcel1,assetExcel2,wb2,
                                    buyFlag,quantFlag,path,buyOrderList,
                                    outerStatusFlag,quantBuyFlag)

                            loopFlag = 1
                            innerStatusFlag = 1

                        #TODO add futures sell call
                        if (sellSignal1 == 1):
                            wbt = xw.Book(path)
                            sheet = wbt.sheets[0]
                            c7 = sheet.range('V'+flagstr).value
                            assetQuant = float(sheet.range('N1').value)
                            
                            quantSellFlag = 20
                            signalSellFlag = 21

                            sellOrderDict[quantSellFlag]=signalSellFlag

                            print(c7)
                            wbt.close()
                            
                            SellOrder(sellFlag,assetQuant,client,coin,c7,ws1,ws2,
                                    flag,assetExcel1,assetExcel2,wb2,path,
                                    sellOrderList,outerStatusFlag,
                                    sellOrderDict,btc_price,signalSellFlag)
                            
                            loopFlag = 1
                            innerStatusFlag = 1
                        
                        if (sellSignal2 == 1):
                            wbt = xw.Book(path)
                            sheet = wbt.sheets[0]
                            c7 = sheet.range('AM'+flagstr).value
                            assetQuant = float(sheet.range('AE1').value)
                            quantSellFlag = 37
                            signalSellFlag = 38

                            sellOrderDict[quantSellFlag]=signalSellFlag

                            print(c7)
                            wbt.close()
                            
                            SellOrder(sellFlag,assetQuant,client,coin,c7,ws1,ws2,
                                    flag,assetExcel1,assetExcel2,wb2,path,
                                    sellOrderList,outerStatusFlag,
                                    sellOrderDict,btc_price,signalSellFlag)
                            
                            loopFlag = 1
                            innerStatusFlag = 1

                        if (sellSignal3 == 1):
                            wbt = xw.Book(path)
                            sheet = wbt.sheets[0]
                            c7 = sheet.range('BD'+flagstr).value
                            assetQuant = float(sheet.range('AV1').value)
                            quantSellFlag = 54
                            signalSellFlag = 55

                            sellOrderDict[quantSellFlag]=signalSellFlag

                            print(c7)
                            wbt.close()
                            
                            SellOrder(sellFlag,assetQuant,client,coin,c7,ws1,ws2,
                                    flag,assetExcel1,assetExcel2,wb2,path,
                                    sellOrderList,outerStatusFlag,
                                    sellOrderDict,btc_price,signalSellFlag)
                            
                            loopFlag = 1
                            innerStatusFlag = 1

                        if (sellSignal4 == 1):
                            wbt = xw.Book(path)
                            sheet = wbt.sheets[0]
                            c7 = sheet.range('BU'+flagstr).value
                            assetQuant = float(sheet.range('BM1').value)
                            quantSellFlag = 71  
                            signalSellFlag = 72

                            sellOrderDict[quantSellFlag]=signalSellFlag

                            print(c7)
                            wbt.close()
                            
                            SellOrder(sellFlag,assetQuant,client,coin,c7,ws1,ws2,
                                    flag,assetExcel1,assetExcel2,wb2,path,
                                    sellOrderList,outerStatusFlag,
                                    sellOrderDict,btc_price,signalSellFlag)
                            
                            loopFlag = 1
                            innerStatusFlag = 1

                        if (sellSignal5 == 1):
                            wbt = xw.Book(path)
                            sheet = wbt.sheets[0]
                            c7 = sheet.range('CL'+flagstr).value
                            assetQuant = float(sheet.range('CD1').value)
                            quantSellFlag = 88
                            signalSellFlag = 89

                            sellOrderDict[quantSellFlag]=signalSellFlag

                            print(c7)
                            wbt.close()
                                                    
                            SellOrder(sellFlag,assetQuant,client,coin,c7,ws1,ws2,
                                    flag,assetExcel1,assetExcel2,wb2,path,
                                    sellOrderList,outerStatusFlag,
                                    sellOrderDict,btc_price,signalSellFlag)
                            
                            loopFlag = 1
                            innerStatusFlag = 1

                        if (innerStatusFlag == 1):
                            if (orderType == "LIMIT"):
                                print ("Все ордера были выставленны, ожидайте 2 минуты для подтверждения")
                                time.sleep(120) 
                            
                            if (orderType == "MARKET"):
                                print ("Все ордера были выставленны, ожидайте 5 секунд для подтверждения")
                                time.sleep(15) 
                            
                            if(len(buyOrderList)>0):
                                BuyOrder(allocFunds,btc_price,client,coin,c6,ws2,
                                        ws1,flag,assetExcel1,assetExcel2,wb2,
                                        buyFlag,quantFlag,path,buyOrderList,
                                        1,quantBuyFlag)
                                
                            if(len(sellOrderList)>0):
                                SellOrder(sellFlag,assetQuant,client,coin,c7,ws1,ws2,
                                        flag,assetExcel1,assetExcel2,wb2,path,
                                        sellOrderList,1,sellOrderDict,btc_price,
                                        signalSellFlag)
                            buyOrderList = []
                            sellOrderList = []
                            sellOrderDict = {}

                        if (loopFlag == 1):
                            break

                        else:
                            print("\n-------------------------------------------------")
                            print (datetime.now(),"Данные были введенны в таблицу, никакого действия не было обнаруженно, бот продолжит работать")
                            print("-------------------------------------------------")
                            wbt.close()
                            break                        

                else:
                    flag = flag + 1
                    
            time.sleep(60) 

        else:
            print("\n-------------------------------------------------")
            print (datetime.now(),"Данное время не соблюдает параметрам, никакого действия не произошло, ждите повторения через одну минуту")
            print("-------------------------------------------------")
            time.sleep(60) 


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Load historical data into csv file
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def LoadTimeStamp():
    url = 'https://api.binance.com/api/v3/klines'
    print ("\nВведите валютную пару: (формат: BTCUSDT)\n")
    symbol = input()

    print ("\nВведите интервал по времени: (формат: 1m,10m,15m,30m,1h,1d,1w)\n")
    interval = input()

    print ("\nВведите дату начала через запятую и с двухзначными числами: (формат 2021.01.30)")
    startDate = input()
    startList = startDate.split(".",2)
    start = str(int(dt.datetime(int(startList[0]),int(startList[1]),int(startList[2])).timestamp()*1000))

    print ("\nВведите дату финала через запятую и с двухзначными числами: (формат 2022.01.30)")
    endDate = input()
    endList = endDate.split(".",2)
    end = str(int(dt.datetime(int(endList[0]),int(endList[1]),int(endList[2])).timestamp()*1000))

    FetchHistData(symbol,interval,start,end,url,endList)


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Call the historical api and construct a data frame
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def FetchHistData(symbol,interval,start,end,url,endList):
    par = {'symbol': symbol, 'interval': interval, 'startTime': start, 'endTime': end, 'limit' : 1000}

    data = pd.DataFrame(json.loads(requests.get(url, params= par).text))
    #format columns name
    data.columns = ['datetime', 'open', 'high', 'low', 'close',
                    'volume','close_time', 'qav', 'num_trades',
                    'taker_base_vol', 'taker_quote_vol', 'ignore']
    
    data.index = [dt.datetime.fromtimestamp(x/1000.0) for x in data.datetime]
    data=data.astype(float)

    limitCheck = str(data.tail(1).index[0])[:-9].split("-",2)

    data.to_csv(symbol+'.csv',mode="a", index = 1, header=True)
    
    if (limitCheck != endList):
        startNewLimit = dt.datetime(int(limitCheck[0]),int(limitCheck[1]),int(limitCheck[2]),
        int(str(data.tail(1).index[0])[+11:-6]),int(str(data.tail(1).index[0])[+14:-3])) + dt.timedelta(minutes=1)
        
        FetchHistData(symbol,interval,str(int(startNewLimit.timestamp()*1000)),end,url,endList)


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Connection to telegram bot for errors
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def TelegramBot(message,path):
    try:
        teleBot = telegram.Bot("1960666049:AAFEaBBvpvNM37i2rCt70JIC1w-Rt1g_v1M")
        teleBot.send_message(-579479570,str(datetime.now())+" "+str(path)+" "+message)
        
    except Exception as e:
        print ("\n Не удалось подключиться к телеграм боту, алго бот продолжит пытаться подключиться")
        print (e)
        time.sleep(5)
        StartBot(path)


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Connection to telegram bot for orders
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def TelegramBotOrder(message,path):
    try:
        teleBot = telegram.Bot("1960666049:AAFEaBBvpvNM37i2rCt70JIC1w-Rt1g_v1M")
        teleBot.send_message(-739688904,str(datetime.now())+" "+str(path)+" "+message)
        
    except Exception as e:
        print ("\n Не удалось подключиться к телеграм боту, алго бот продолжит пытаться подключиться")
        print (e)
        time.sleep(5)
        StartBot(path)


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Starter
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def StartBot(path):
            
    try:
        MainFunc(path)

    except Exception as e:
        if ("timed out" in str(e)):
            print(datetime.now(),"Произошла ошибка во время подключения программа попробует еще раз через 5 секунд")

            TelegramBot("Произошла ошибка во время подключения бот будет продолжать пытаться подключиться"
                + str(e),path)

            time.sleep(5) 
            StartBot(path)
            
        elif ("Timestamp for this request was" in str(e)):
            print(datetime.now(),
                "Произошла ошибка: Синхронизируйте время вашей системы в настройках. Программа попробует еще раз через 5 секунд")
                        
            TelegramBot("Произошла ошибка: Синхронизируйте время вашей системы в настройках бот будет продолжать пытаться подключиться"
                + str(e),path)
            time.sleep(5) 
            StartBot(path)

        elif ("Timestamp for this request is" in str(e)):
            print(datetime.now(),
                "Произошла ошибка: Синхронизируйте время вашей системы в настройках. Программа попробует еще раз через 5 секунд")
                        
            TelegramBot("Произошла ошибка: Синхронизируйте время вашей системы в настройках бот будет продолжать пытаться подключиться"
                + str(e),path)
            time.sleep(5) 
            StartBot(path)

        elif ("Max retries exceeded with url" in str(e)):
                print(datetime.now(),
                    "Произошла ошибка во время подключения программа попробует еще раз через 5 секунд")

                TelegramBot("Произошла ошибка во время подключения бот будет продолжать пытаться подключиться"
                    + str(e),path)

                time.sleep(5) 
                StartBot(path)

        elif ("ConnectionResetError" in str(e)):
            print(datetime.now(),
                "Произошла ошибка во время подключения программа попробует еще раз через 5 секунд")

            TelegramBot("Произошла ошибка во время подключения бот будет продолжать пытаться подключиться"
                + str(e),path)

            time.sleep(5) 
            StartBot(path)

        else:
            print("\n-------------------------------------------------")
            print("Произошла ошибка:")
                        
            TelegramBot("Произошла ошибка бот был остановлен. Код ошибки:" + str(e),path)

            print(e)
            print("-------------------------------------------------")
            Menu()


#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Display menu and process choice
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def Menu():
    print("\n\n-------------------------------------------------")
    print ("Выберите функцию:")
    print ("0 - Проверить баланс")
    print ("1 - Начать торговлю с начала")
    print ("2 - Продолжить торговлю")
    print ("3 - Продать валюту")
    print ("4 - Произвести выгрузку")
    print ("5 - Проверить Телеграм бот")
    print ("9 - Выйти из приложения\n")
    choice = input()

    if (choice == "1"):
        try:
            ClearExcel()
            StartBot(GetPath())
            
        except Exception as e:
            print("\n-------------------------------------------------")
            print("Произошла ошибка:")

            TelegramBot("Произошла ошибка бот был остановлен. Код ошибки:" + str(e),path)

            print(e)
            print("-------------------------------------------------")
            Menu()

    elif (choice == "2"):
        StartBot(GetPath())

    elif (choice == "3"):
        try:
            SellFunc(GetPath())
            
        except Exception as e:
            if ("timed out" in str(e)):
                print(datetime.now(),"Произошла ошибка во время подключения попробуйте еще раз")

                Menu()
                
            elif ("Timestamp for this request was" in str(e)):
                print(datetime.now(),
                    "Произошла ошибка: Синхронизируйте время вашей системы в настройках. Попробуйте еще раз")

                Menu()

            else:
                print("\n-------------------------------------------------")
                print("Произошла ошибка:")
                print(e)
                print("-------------------------------------------------")
                Menu()
                
    elif (choice == "4"):
        try:
            LoadTimeStamp()
            
        except Exception as e:
            print("\n-------------------------------------------------")
            print("Произошла ошибка:")

            print(e)
            print("-------------------------------------------------")
            Menu()
            
    elif (choice == "5"):
        try:
            TelegramBot("Произошла проверка Бота",path)
            
        except Exception as e:
            print("\n-------------------------------------------------")
            print("Произошла ошибка:")

            print(e)
            print("-------------------------------------------------")
            Menu()

    elif (choice == "0"):
        CheckBalance()
        Menu()

    elif (choice == "9"):
        wb = load_workbook(GetPath())
        wb.save(GetPath())
        wb.close()

        wbx = xw.Book((GetPath()))
        wbx.save(GetPath())
        wbx.close()

        for proc in psutil.process_iter():
            if proc.name() == "EXCEL.EXE":
                proc.kill()
            if proc.name() == "excel.exe":
                proc.kill()     
                   
        sys.exit()

    else:
        print("\nВыбранный вами режим не существует попробуйте еще раз")
        Menu()
        

#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#! Init
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
Menu()
